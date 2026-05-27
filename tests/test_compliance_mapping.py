"""Tests for compliance framework mapping (Group B)."""
import csv
import io
import os
import sys
import tempfile
from pathlib import Path

import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from utils.compliance import (
    map_vuln,
    enrich_vulnerabilities,
    load_frameworks,
    FRAMEWORK_KEYS,
)
from utils.compliance.mapper import _cache as MAPPER_CACHE
from utils.exports.csv_builder import build_csv


SAMPLE_VULNS = [
    {
        "type": "SQL Injection",
        "severity": "critical",
        "url": "https://example.com/login",
        "parameter": "user",
        "payload": "'",
        "evidence": "syntax error",
        "description": "sqli",
        "remediation": "parameterize",
        "cve_references": [],
        "cvss_score": 9.8,
    },
    {
        "type": "XSS",
        "severity": "high",
        "url": "https://example.com/search",
        "parameter": "q",
        "payload": "<script>",
        "evidence": "reflected",
        "description": "xss",
        "remediation": "encode",
        "cve_references": [],
        "cvss_score": 7.5,
    },
    {
        "type": "Made Up Vuln Type",
        "severity": "low",
        "url": "https://example.com/x",
        "parameter": "",
        "payload": "",
        "evidence": "",
        "description": "",
        "remediation": "",
        "cve_references": [],
        "cvss_score": 1.0,
    },
]


@pytest.fixture(autouse=True)
def clear_cache():
    """Clear mapper cache between tests."""
    MAPPER_CACHE.clear()
    yield
    MAPPER_CACHE.clear()


class TestFrameworkLoading:
    def test_load_all(self):
        frameworks = load_frameworks()
        assert "PCI-DSS" in frameworks
        assert "SOC 2" in frameworks
        assert "ISO 27001" in frameworks

    def test_load_subset(self):
        frameworks = load_frameworks(["pci_dss"])
        assert list(frameworks.keys()) == ["PCI-DSS"]

    def test_load_unknown_key(self):
        frameworks = load_frameworks(["invalid_key"])
        assert frameworks == {}

    def test_framework_schema(self):
        frameworks = load_frameworks(["pci_dss"])
        pci = frameworks["PCI-DSS"]
        assert pci["framework"] == "PCI-DSS"
        assert pci["version"] == "4.0"
        assert "controls" in pci
        assert "vuln_mappings" in pci
        assert "6.2.4" in pci["controls"]
        assert "SQL Injection" in pci["vuln_mappings"]


class TestMapping:
    def test_known_vuln(self):
        result = map_vuln("SQL Injection")
        assert "PCI-DSS" in result
        assert "SOC 2" in result
        assert "ISO 27001" in result
        assert len(result["PCI-DSS"]) > 0
        ctrl = result["PCI-DSS"][0]
        assert "control_id" in ctrl
        assert "title" in ctrl
        assert "category" in ctrl

    def test_unknown_vuln(self):
        result = map_vuln("Some Made Up Vuln")
        assert result["PCI-DSS"] == []
        assert result["SOC 2"] == []
        assert result["ISO 27001"] == []

    def test_framework_filter(self):
        result = map_vuln("XSS", framework_keys=["pci_dss"])
        assert "PCI-DSS" in result
        assert "SOC 2" not in result
        assert "ISO 27001" not in result


class TestEnrichment:
    def test_enrich_in_place(self):
        vulns = [dict(v) for v in SAMPLE_VULNS]
        enrich_vulnerabilities(vulns)
        for v in vulns:
            assert "compliance" in v
            assert "PCI-DSS" in v["compliance"]
            assert "SOC 2" in v["compliance"]
            assert "ISO 27001" in v["compliance"]

    def test_enrich_unknown_vuln(self):
        vulns = [dict(SAMPLE_VULNS[2])]
        enrich_vulnerabilities(vulns)
        assert vulns[0]["compliance"]["PCI-DSS"] == []

    def test_enrich_with_filter(self):
        vulns = [dict(SAMPLE_VULNS[0])]
        enrich_vulnerabilities(vulns, framework_keys=["soc2"])
        assert "SOC 2" in vulns[0]["compliance"]
        assert "PCI-DSS" not in vulns[0]["compliance"]


class TestCSVCompliance:
    def test_csv_with_compliance(self):
        vulns = [dict(v) for v in SAMPLE_VULNS]
        enrich_vulnerabilities(vulns)
        text = build_csv({"vulnerabilities": vulns})
        body = text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(body))
        header = next(reader)
        assert "compliance_pci_dss" in header
        assert "compliance_soc2" in header
        assert "compliance_iso_27001" in header
        rows = list(reader)
        assert len(rows) == 3
        pci_idx = header.index("compliance_pci_dss")
        assert "6.2.4" in rows[0][pci_idx]

    def test_csv_without_compliance(self):
        text = build_csv({"vulnerabilities": SAMPLE_VULNS})
        body = text.lstrip("\ufeff")
        reader = csv.reader(io.StringIO(body))
        header = next(reader)
        assert "compliance_pci_dss" not in header


class TestXLSXCompliance:
    def test_compliance_sheet_populated(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        from utils.exports.xlsx_builder import build_xlsx

        vulns = [dict(v) for v in SAMPLE_VULNS]
        enrich_vulnerabilities(vulns)
        results = {
            "target": "https://example.com",
            "duration": 1.0,
            "vulnerabilities": vulns,
            "severity_summary": {},
            "urls_crawled": 1,
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            path = tf.name

        success = build_xlsx(results, path, interactive=False)
        assert success

        from openpyxl import load_workbook
        wb = load_workbook(path)
        compliance = wb["Compliance"]
        assert compliance.max_row > 1
        all_values = []
        for row in compliance.iter_rows(min_row=2, values_only=True):
            all_values.extend([str(c) for c in row])
        assert any("6.2.4" in v for v in all_values)
        assert "Control Summary" in wb.sheetnames

        os.unlink(path)

    def test_control_summary_sheet(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        from utils.exports.xlsx_builder import build_xlsx

        vulns = [dict(v) for v in SAMPLE_VULNS]
        enrich_vulnerabilities(vulns)
        results = {
            "target": "x",
            "duration": 1,
            "vulnerabilities": vulns,
            "severity_summary": {},
            "urls_crawled": 0,
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            path = tf.name

        build_xlsx(results, path, interactive=False)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        cs = wb["Control Summary"]
        headers = [c.value for c in cs[1]]
        assert headers == ["framework", "control_id", "control_title", "vuln_count", "severity_max"]
        assert cs.max_row > 1

        os.unlink(path)

    def test_no_control_summary_when_disabled(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            pytest.skip("openpyxl not installed")

        from utils.exports.xlsx_builder import build_xlsx

        results = {
            "target": "x",
            "duration": 1,
            "vulnerabilities": SAMPLE_VULNS,
            "severity_summary": {},
            "urls_crawled": 0,
        }

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            path = tf.name

        build_xlsx(results, path, interactive=False)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert "Compliance" in wb.sheetnames
        assert "Control Summary" not in wb.sheetnames

        os.unlink(path)
